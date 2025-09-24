import streamlit as st
import pdfplumber
import fitz
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="pdfminer")
import docx
import re
import os
import google.generativeai as genai
import csv
import json
from openpyxl import Workbook, load_workbook

# =======================================================================
# ===== CONFIG
# =======================================================================

# Dùng st.secrets để quản lý khóa API một cách an toàn hơn
GEMINI_API_KEY = "AIzaSyCSenmJGRf2VJ9WId1SwQpfL3dMRRaHWmw"
# GEMINI_API_KEY = "AIzaSyABwa4KRue_M2A7l2YHAN4J2tPQJ5s33Ig"
# GEMINI_API_KEY = "AIzaSyBGSw2-NoZXd3HT_jWK1HoNzX7WhHcaBNA"
genai.configure(api_key=GEMINI_API_KEY)

# =======================================================================
# ===== FUNCTIONS
# =======================================================================

rubric = """
Tiêu chí chấm điểm thuộc lĩnh vực Công nghệ thông tin. Tổng điểm tối đa là 10 điểm, chia thành 5 nhóm tiêu chí chính:
1. Nội dung và kiến thức chuyên môn (4.5 điểm)
2. Cấu trúc và tổ chức bài viết (2 điểm)
3. Trình bày và diễn đạt (1.5 điểm)
4. Phân tích, lập luận và giải thích (1.5 điểm)
5. Tài liệu tham khảo và trích dẫn (0.5 điểm)
Ngoài ra có tiêu chí trừ điểm nếu bài làm vi phạm yêu cầu nghiêm trọng.

Tiêu chí 1 – Nội dung và kiến thức chuyên môn (4.5 điểm):
• Đáp ứng yêu cầu đề bài (1.5 điểm): bài làm đúng trọng tâm, không lạc đề, bao quát đủ các khía cạnh.
• Tính chính xác kiến thức (2 điểm): nội dung đúng, không sai nghiêm trọng về lý thuyết, thuật toán, phương pháp.
• Tính sáng tạo và thực tiễn (1 điểm): có tiếp cận mới, vận dụng thực tế, có dẫn chứng rõ.

Tiêu chí 2 – Cấu trúc và tổ chức bài viết (2 điểm):
• Cấu trúc (1.2 điểm): bài gồm mở bài, thân bài, kết luận; bố cục rõ ràng, không rời rạc.
• Logic liên kết (0.8 điểm): các ý nối liền mạch, có sử dụng từ nối để giữ tính liên tục.

Tiêu chí 3 – Trình bày và diễn đạt (1.5 điểm):
• Ngôn ngữ chuyên ngành (0.6 điểm): sử dụng đúng thuật ngữ CNTT, tránh cách nói đơn giản hóa.
• Chính tả, ngữ pháp (0.6 điểm): không mắc lỗi nghiêm trọng, diễn đạt mạch lạc.
• Hình thức trình bày (0.3 điểm): rõ ràng, khoa học, có số trang/mục lục nếu cần.

Tiêu chí 4 – Phân tích, lập luận và giải thích (1.5 điểm):
• Lập luận chặt chẽ (0.7 điểm): có cơ sở rõ ràng, giải thích hợp lý.
• Ví dụ, dẫn chứng (0.5 điểm): có minh họa bằng thực tế, sơ đồ hoặc thuật toán.
• Phản biện, đánh giá (0.3 điểm): đưa nhiều góc nhìn, so sánh giải pháp khác nhau.

Tiêu chí 5 – Tài liệu tham khảo và trích dẫn (0.5 điểm):
• Tài liệu chất lượng (0.3 điểm): có nguồn gốc rõ ràng, chính thống.
• Trích dẫn đúng chuẩn (0.2 điểm): theo quy chuẩn như APA, IEEE, Harvard...

Tiêu chí trừ điểm:
• Lạc đề hoặc sai nghiêm trọng về kiến thức: trừ toàn bộ điểm phần nội dung.
• Trình bày quá sơ sài, không có cấu trúc rõ ràng: trừ tối đa 2 điểm.
• Viết sai chính tả, ngữ pháp quá nhiều: trừ tối đa 1 điểm.
• Không trích nguồn nhưng dùng tài liệu ngoài: trừ 0.5 điểm.
"""

def read_docx(file):
    """Đọc nội dung từ file .docx."""
    doc = docx.Document(file)
    full_text = []
    for para in doc.paragraphs:
        full_text.append(para.text)
    return '\n'.join(full_text)

def read_pdf(file):
    """Đọc nội dung từ file .pdf."""
    text = ""
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            text += page.extract_text(x_tolerance=2) + "\n\n"
    return text

def read_txt(file):
    """Đọc nội dung từ file .txt."""
    return file.getvalue().decode('utf-8')

def read_file(file):
    """Xác định loại file và đọc nội dung tương ứng."""
    file_type = file.name.split('.')[-1].lower()
    if file_type == 'pdf':
        return read_pdf(file)
    elif file_type == 'docx':
        return read_docx(file)
    elif file_type == 'txt':
        return read_txt(file)
    else:
        raise ValueError("Định dạng file không được hỗ trợ.")

def save_result_to_excel(course, filename, essay_text, score, model_name, file_path="grading_results_norag.xlsx"):
    """Lưu kết quả chấm điểm vào file Excel."""
    essay_preview = essay_text[:300] + ("..." if len(essay_text) > 300 else "")
    
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Tên môn học", "Tên file", "Nội dung bài luận", "Mô hình", "Điểm Không RAG"])

    ws.append([course, filename, essay_preview, model_name, score])
    wb.save(file_path)

def grade_essay(essay_text, course_context, student_context, model_name="gemini-2.0-flash"):
    """
    Sử dụng LLM để chấm điểm bài luận có thêm ngữ cảnh.
    """
    temperature_value = 0.7
    model = genai.GenerativeModel(model_name)
    
    # Prompt mới, được thêm ngữ cảnh từ người dùng
    prompt = f"""
    Bạn là một tiến sĩ đại học có nhiều năm kinh nghiệm trong lĩnh vực Công nghệ thông tin. Bạn cực kỳ nghiêm khắc và chuyên nghiệp trong việc chấm điểm, luôn đưa ra phản hồi chi tiết, khách quan và có tính xây dựng.

    Hãy phân tích và chấm điểm bài luận của sinh viên dưới đây dựa trên bộ tiêu chí chi tiết mà tôi cung cấp. Bài luận này thuộc môn học {course_context}.

    Bộ tiêu chí chấm điểm:
    {rubric}

    Yêu cầu cụ thể:
    1.  Chấm điểm từng tiêu chí: Phân tích kỹ lưỡng và đưa ra điểm số cụ thể (có thể là số thập phân) cho từng tiêu chí trong bộ rubric. Sử dụng thang điểm từ 0 đến 10 một cách linh hoạt, không giới hạn điểm trong một khoảng hẹp.
    2.  Nhận xét chi tiết:
            Điểm mạnh: Nêu rõ những mặt tích cực mà sinh viên đã làm được, ví dụ: "Kiến thức chính xác và lập luận chặt chẽ."
            Điểm cần cải thiện: Đề xuất những điểm mà sinh viên có thể cải thiện trong bài viết tiếp theo để đạt điểm cao hơn.
    3.  Tính tổng điểm cuối cùng: Tổng hợp điểm từ các tiêu chí và đưa ra tổng điểm cuối cùng trên thang điểm 10.
    4.  Thể hiện ngữ cảnh sinh viên: Dựa vào mô tả {student_context}, hãy điều chỉnh mức độ nghiêm khắc khi chấm điểm. Ví dụ, với sinh viên năm nhất, tập trung vào cấu trúc cơ bản; với sinh viên năm 4, yêu cầu cao hơn về mặt học thuật và phân tích.

    Định dạng đầu ra phải là một chuỗi JSON duy nhất, không thêm bất kỳ văn bản nào khác.

    Bài luận cần chấm:
    ---
    {essay_text}
    ---

    Định dạng đầu ra:
    Hãy trả về kết quả theo cấu trúc sau, đảm bảo mọi thông tin đều được trình bày rõ ràng và dễ đọc.
    -   Điểm tổng: [Điểm số linh hoạt từ 0-10]
    -   Điểm mạnh: [Đưa ra ít nhất 3 điểm mạnh nổi bật]
    -   Điểm cần cải thiện: [Đưa ra ít nhất 3 điểm cần cải thiện cụ thể]
    -   Nhận xét: [Phản hồi chi tiết theo yêu cầu trên]
    
    Đầu ra JSON:
    """

    try:
        response = model.generate_content(prompt, generation_config=genai.types.GenerationConfig(temperature=temperature_value))
        
        json_output = response.text
        match = re.search(r'```json(.*?)```', json_output, re.DOTALL)
        if match:
            json_str = match.group(1).strip()
            result_data = json.loads(json_str)
        else:
            result_data = json.loads(json_output)
            
        score = int(result_data.get('Điểm tổng', 0))
        general_comment = result_data.get('Nhận xét chung', '')
        strengths = result_data.get('Điểm mạnh', [])
        improvements = result_data.get('Điểm cần cải thiện', [])
        
        result_text = f"Nhận xét chung: {general_comment}\n\n"
        
        if strengths:
            result_text += "Điểm mạnh:\n" + "\n".join([f"- {s}" for s in strengths]) + "\n\n"
        if improvements:
            result_text += "Điểm cần cải thiện:\n" + "\n".join([f"- {i}" for i in improvements])
            
        return score, result_text
    
    except Exception as e:
        return 0, f"Có lỗi xảy ra trong quá trình chấm điểm: {e}"

# =======================================================================
# ===== STREAMLIT UI
# =======================================================================

st.set_page_config(page_title="Chấm điểm bài luận", page_icon="📄", layout="wide")

st.markdown("""
    <style>
    .main {
        background-color: #f9fafc;
    }
    .st-emotion-cache-1wv936z {
        border-radius: 15px;
        padding: 20px;
        background-color: #ffffff;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        border: 1px solid #e0e0e0;
    }
    .title-container {
        display: flex;
        align-items: center;
        gap: 15px;
    }
    .score-box {
        background-color: #f0f2f6;
        border-radius: 10px;
        padding: 10px 20px;
        text-align: center;
        border: 1px dashed #d1d5db;
        margin-top: 10px;
    }
    .score-text {
        font-size: 3rem;
        font-weight: bold;
        color: #1a73e8;
    }
    </style>
""", unsafe_allow_html=True)

st.title("📄 Hệ thống chấm điểm bài luận tự động")

# Khai báo model_name ở đây
model_name = "gemini-2.0-flash"

if 'score' not in st.session_state:
    st.session_state.score = None
if 'result_text' not in st.session_state:
    st.session_state.result_text = None

col1, col2 = st.columns([1, 2])

with col1:
    st.subheader("📤 Tải lên bài luận")
    uploaded_file = st.file_uploader("Tải lên file (PDF, DOCX, TXT)", type=["pdf", "docx", "txt"])

    course_context = st.selectbox("📘 Môn học", ["Chiến lược kinh doanh thương mại điện tử", "Hệ thống thanh toán điện tử", "Marketing kỹ thuật số", "Phát triển ứng dụng thương mại điện tử", "Quản trị dự án thương mại điện tử", "Thương mại điện tử"])
    student_context = st.selectbox("🎓 Ngữ cảnh sinh viên", [
        "Năm 4: Sinh viên cần đạt chuẩn luận văn tốt nghiệp: viết học thuật hoàn chỉnh, có đặt vấn đề, cơ sở lý thuyết, phương pháp, phân tích, kết quả, kết luận, sử dụng trích dẫn chuẩn, thể hiện tư duy nghiên cứu độc lập và đóng góp mới.", 
        "Năm 3: Sinh viên phải thể hiện lập luận chặt chẽ hơn, biết sử dụng tài liệu tham khảo (trích dẫn đúng cách), trình bày theo chuẩn học thuật, có phân tích, so sánh, đánh giá thay vì chỉ mô tả.", 
        "Năm 2: Sinh viên bắt đầu học kỹ năng viết nâng cao hơn. Cần có cấu trúc 3 phần (mở bài, thân bài, kết luận), biết triển khai luận điểm theo đoạn văn mạch lạc, có ví dụ minh họa cơ bản.", 
        "Năm 1: Sinh viên chỉ mới làm quen viết luận. Đánh giá chủ yếu ở sự rõ ràng, logic cơ bản, cách trình bày ý tưởng. Không yêu cầu nhiều về trích dẫn học thuật hay cấu trúc phức tạp."
    ])

    if st.button("🚀 Chấm điểm"):
        if uploaded_file is None:
            st.warning("Vui lòng tải lên một file bài luận.")
        else:
            with st.status("🚀 Đang chấm điểm...", expanded=True) as status_box:
                status_box.write("1. Đang đọc file...")
                try:
                    essay_text = read_file(uploaded_file)
                except Exception as e:
                    status_box.update(label=f"❌ Lỗi: Không thể đọc file. {e}", state="error", expanded=False)
                    st.error("Có lỗi xảy ra khi đọc file. Vui lòng thử lại.")
                    st.stop()
                
                status_box.write("2. Đang phân tích bài luận và gọi mô hình AI...")
                
                # Gọi hàm chấm điểm và truyền thêm ngữ cảnh
                score_value, result_text = grade_essay(
                    essay_text=essay_text, 
                    course_context=course_context, 
                    student_context=student_context, 
                    model_name=model_name
                )
                
                status_box.update(label=f"✅ Đã hoàn tất! Kết quả: {score_value} điểm.", state="complete", expanded=False)
                
                try:
                    save_result_to_excel(course_context, uploaded_file.name, essay_text, score_value, model_name)
                    st.success(f"✅ Kết quả đã được lưu vào grading_results_norag.xlsx (Điểm Không RAG: {score_value})")
                except Exception as e:
                    st.error(f"❌ Lỗi khi lưu vào Excel: {e}")

                st.session_state.score = score_value
                st.session_state.result_text = result_text

with col2:
    st.header("📊 Kết quả chấm điểm")
    if st.session_state.score is not None and st.session_state.result_text is not None:
        score_value = st.session_state.score
        result_text = st.session_state.result_text
        
        st.markdown(f"""
            <div class="score-box">
                <div style="color: #000">Điểm tổng</div>
                <div class="score-text">{score_value}</div>
            </div>
        """, unsafe_allow_html=True)

        st.subheader("📝 Nhận xét chi tiết")
        st.write(result_text)
    else:
        st.info("👉 Kết quả sẽ hiển thị tại đây sau khi chấm điểm.")